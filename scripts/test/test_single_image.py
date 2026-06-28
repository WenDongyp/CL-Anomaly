import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

from llava.model.builder import load_pretrained_model
from llava.mm_utils import get_model_name_from_path, process_images, tokenizer_image_token
from llava.constants import IMAGE_TOKEN_INDEX, DEFAULT_IMAGE_TOKEN, DEFAULT_IM_START_TOKEN, DEFAULT_IM_END_TOKEN, IGNORE_INDEX
from llava.conversation import conv_templates, SeparatorStyle
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score

from PIL import Image
import requests
import copy
import torch

import json
import time
import os

import sys
import warnings

warnings.filterwarnings("ignore")

import argparse
from openpyxl import load_workbook, Workbook

def cal_metrics(gt_data, responses, model_path, data_bench):

    class_name = set()
    for item in gt_data:
        class_name.add(item['image'].split('/')[1])

    class_name = list(class_name)

    acc_results = []
    for cls in class_name:
        y_true = []
        y_pred = []
        for index, item in enumerate(gt_data):
            if item['image'].split('/')[1] == cls:
                anomaly = 1 if item['metadata']['anomaly'] else 0
                #print(anomaly)
                y_true.append(anomaly)
                pred = 1 if 'Yes' in responses[index] or 'yes' in responses[index] else 0
                y_pred.append(pred)

        acc = accuracy_score(y_true, y_pred)
        acc_results.append(acc)
        print(f'Class: {cls}, Accuracy: {acc:.4f}')

    print(f'Mean Accuracy: {sum(acc_results) / len(acc_results):.4f}')
    avg_acc = sum(acc_results) / len(acc_results)

    result_dir = '/data1/dongwen_data/CL-Anomaly/results'
    os.makedirs(result_dir, exist_ok=True)
    result_file = os.path.join(result_dir, 'CL-Anomaly-results.xlsx')

    if os.path.isfile(result_file):
        wb = load_workbook(result_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['model_path', 'data_json', 'Accuracy'])

    ws.append([
        model_path,
        data_bench,
        round(avg_acc, 4)
    ])
    wb.save(result_file)
    wb.close()

def eval_model(args):
    #path = "/data1/dongwen_data/Model/llava-onevision-qwen2-7b-ov"
    #emb = torch.load(os.path.join(path, "model-00001-of-00004.safetensors"), map_location="cpu")
    #print(emb["model.embed_tokens.weight"].shape)

    model_path = args.model_checkpoint
    model_base = args.model_base
    #model_name = get_model_name_from_path(model_path)
    model_name = "llava_qwen_lora"
    device = "cuda"
    #device_map = "auto"
    device_map = "cuda"
    overwrite_config = {'tie_word_embeddings': False, 'use_cache': True, 'vocab_size': 152064}
    #overwrite_config = {'vocab_size': 152064}  #7b 151936
    print("model_path", model_path)
    print("model_base", model_base)
    print("model_name", model_name)
    print("data bench", args.bench_json)


    tokenizer, model, image_processor, max_length = load_pretrained_model(model_path, model_base, model_name, device_map=device_map, cache_dir='./cache', torch_dtype="bfloat16", overwrite_config=overwrite_config)

    if args.size != '7b':
        model.lm_head.weight = model.model.embed_tokens.weight
        print("Testing 0.5B model, set lm_head weight to embed_tokens weight")
    else:
        print("Testing 7B model, no need to set lm_head weight")

    model.eval()

    responses = []

    with open(os.path.join(args.data_dir, args.bench_json), 'r') as f:
        data = json.load(f)

    for index, d in enumerate(data):
        image_path = os.path.join(args.data_dir, d['image'])
        image = Image.open(image_path).convert("RGB")
        # if the longest side of the image is greater than 1024, resize it to 1024 while keeping the aspect ratio
        if max(image.size) > 1024:
            if image.width > image.height:
                new_width = 1024
                new_height = int(1024 * image.height / image.width)
            else:
                new_height = 1024
                new_width = int(1024 * image.width / image.height)
            image = image.resize((new_width, new_height))

        image_tensor = process_images([image], image_processor, model.config)
        image_tensor = [_image.to(dtype=torch.bfloat16, device=device) for _image in image_tensor]

        conv_template = "qwen_1_5"  # Make sure you use correct chat template for different models
        #question = DEFAULT_IMAGE_TOKEN + "\nAre there any defects for the object in the image? "
        question = DEFAULT_IMAGE_TOKEN + "\nAre there any defects for the object in the image? Please reply with only one word 'Yes' or 'No'. "
        # question = DEFAULT_IMAGE_TOKEN + "\nDo you notice any anomalies in the image? Please reply with 'Yes' or 'No'."
        #print(question)
        conv = copy.deepcopy(conv_templates[conv_template])
        conv.append_message(conv.roles[0], question)
        conv.append_message(conv.roles[1], None)
        prompt_question = conv.get_prompt()

        input_ids = tokenizer_image_token(prompt_question, tokenizer, IMAGE_TOKEN_INDEX, return_tensors="pt").unsqueeze(0).to(device)
        image_sizes = [image.size]

        cont = model.generate(
            input_ids,
            images=image_tensor,
            image_sizes=image_sizes,
            do_sample=False,
            temperature=0,
            max_new_tokens=4096,
        )
        text_outputs = tokenizer.batch_decode(cont, skip_special_tokens=True)
        #print(index, text_outputs[0])
        #print(f"Index: {index}")
        #print(f"Question: {question}")
        #print(f"Output: {text_outputs[0]}")
        #print("-" * 50)
        responses.append(text_outputs[0])

    cal_metrics(data, responses, model_path, args.bench_json)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Test detection performance of the model")
    parser.add_argument("--data_dir", type=str, default='your_data_dir', help="Path to your data directory")
    parser.add_argument("--bench_json", type=str, default='VisA/test_data.json', help="Path to your benchmark json file")
    parser.add_argument("--model_checkpoint", type=str, default='your_model_path', help="Path to your pretrained model")
    parser.add_argument("--model_base", type=str, default='your_base_model_path', help="Path to your pretrained model")
    parser.add_argument("--size", type=str, default='7b', help="Model size")
    parser.add_argument("--save_path", type=str, default='your_save_path', help="Path to save the results")
    args = parser.parse_args()

    eval_model(args)


